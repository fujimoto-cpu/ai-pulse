"""
スプシ「選択肢」タブ用のXLSXを生成。
全質問の選択肢を1つのタブで統合管理する。

カラム構造：
  質問キー | 値 | 絵文字 | ラベル | 説明 | スコア | 並び順 | 有効フラグ

出力：~/Downloads/KONNEKT_AIPulse_選択肢.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

wb = Workbook()
ws = wb.active
ws.title = "選択肢"

# ヘッダー
headers = ["質問キー", "値", "絵文字", "ラベル", "説明", "スコア", "並び順", "有効フラグ"]
ws.append(headers)
for cell in ws[1]:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN

# データ
choices = [
    # Q1 使用頻度
    ("q1", "ほぼ毎日", "🔥", "ほぼ毎日", "", 10, 1, True),
    ("q1", "週2-3日", "⚡", "週2〜3日", "", 5, 2, True),
    ("q1", "週1日", "🌱", "週1日くらい", "", 2, 3, True),
    ("q1", "ほぼ使ってない", "😴", "ほぼ使ってない", "", 0, 4, True),
    # Q2 場面（複数選択・スコアは1個1XPで上限6）
    ("q2", "メール作成", "✉️", "メール作成", "", 1, 1, True),
    ("q2", "議事録", "📝", "議事録", "", 1, 2, True),
    ("q2", "リサーチ", "🔍", "リサーチ", "", 1, 3, True),
    ("q2", "企画書", "📋", "企画書", "", 1, 4, True),
    ("q2", "デザイン制作", "🎨", "デザイン制作", "", 1, 5, True),
    ("q2", "コピー作成", "✍️", "コピー作成", "", 1, 6, True),
    ("q2", "データ集計", "📊", "データ集計", "", 1, 7, True),
    ("q2", "書類処理", "📄", "書類処理", "", 1, 8, True),
    # Q4 成長実感
    ("q4", "😫", "😫", "後退", "先月より使えてない／むしろ離れた", 0, 1, True),
    ("q4", "😐", "😐", "変化なし", "先月と同じくらい", 1, 2, True),
    ("q4", "😊", "😊", "少し成長", "新しい使い方を1〜2個身につけた", 2, 3, True),
    ("q4", "🔥", "🔥", "急成長", "自分でも驚くくらい使いこなしてる", 3, 4, True),
    # Q5 来月の方向性
    ("q5", "個人向上", "📚", "個人向上（自分のスキル磨く）", "", 2, 1, True),
    ("q5", "チーム巻き込み", "🤝", "チーム巻き込み（広めていく）", "", 2, 2, True),
    ("q5", "新ツール開拓", "🆕", "新ツール開拓(新しい挑戦)", "", 2, 3, True),
    ("q5", "現状維持", "🌿", "現状維持でいい", "", 1, 4, True),
    # Q6 チーム共有
    ("q6", "😫", "😫", "独占中", "誰にも話してない／聞かれてない", 0, 1, True),
    ("q6", "😐", "😐", "チラ話", "雑談レベルで個別にちょっと", 1, 2, True),
    ("q6", "😊", "😊", "MTGで共有", "チームMTGでデモ・共有した", 2, 3, True),
    ("q6", "🔥", "🔥", "展開リード", "ベスプラとして全社に広めてる", 2, 4, True),
    # できたこと（タグ・複数選択）
    ("didable", "多言語対応", "🌍", "多言語対応（英語・中国語のメール作成）", "", 0, 1, True),
    ("didable", "データ分析", "📊", "データ分析・レポート作成", "", 0, 2, True),
    ("didable", "企画完結", "📝", "企画書・提案書を一人で完結", "", 0, 3, True),
    ("didable", "広範リサーチ", "🔍", "広範囲・短時間リサーチ", "", 0, 4, True),
    ("didable", "デザイン高速化", "🎨", "デザイン初稿の高速化", "", 0, 5, True),
    ("didable", "自動化", "🤖", "繰り返し作業の自動化", "", 0, 6, True),
    ("didable", "議事録要約", "💬", "議事録・要約の自動化", "", 0, 7, True),
    ("didable", "コード活用", "💻", "コード・スクリプト作成", "", 0, 8, True),
    # 困り度
    ("trouble", "😌", "😌", "全然困ってない", "スムーズに使えてる", 0, 1, True),
    ("trouble", "🙂", "🙂", "少し困ってる", "たまに詰まる程度", 0, 2, True),
    ("trouble", "😐", "😐", "中程度", "ちょこちょこ困りごとあり", 0, 3, True),
    ("trouble", "😟", "😟", "かなり困ってる", "やりたいことができない", 0, 4, True),
    ("trouble", "😣", "😣", "非常に困ってる", "AI使うこと自体が大変", 0, 5, True),
]

for c in choices:
    ws.append(c)

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 30
ws.column_dimensions["E"].width = 40
ws.column_dimensions["F"].width = 8
ws.column_dimensions["G"].width = 8
ws.column_dimensions["H"].width = 12

out = Path.home() / "Downloads" / "KONNEKT_AIPulse_選択肢.xlsx"
wb.save(out)
print(f"✅ 保存: {out}")
print(f"   選択肢数: {len(choices)}")
