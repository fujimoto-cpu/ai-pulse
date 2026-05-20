"""
KONNEKT_AIPulse_データ.xlsx を生成するスクリプト。
出力先: ~/Downloads/KONNEKT_AIPulse_データ.xlsx

使い方:
  python3 scripts/create_sheet.py

生成後の手順:
  1. Google Drive (drive.google.com) を開く
  2. ~/Downloads/KONNEKT_AIPulse_データ.xlsx をドラッグ＆ドロップ
  3. ファイルを右クリック → 「アプリで開く」→「Google スプレッドシート」
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def style_header(ws, last_col):
    """1行目をヘッダーとして装飾。"""
    for cell in ws[1][:last_col]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


wb = Workbook()

# ============ タブ1: 名簿 ============
ws1 = wb.active
ws1.title = "名簿"
ws1.append(["名前", "部署", "メアド", "有効フラグ", "PIN"])

members = [
    ["由羽 弘明", "経営・営業", "", True, "1111"],
    ["小柳 雅裕", "経営・営業", "", True, "1111"],
    ["鈴木 由洋", "経営・営業", "", True, "1111"],
    ["山﨑 丈嗣", "経営・営業", "", True, "1111"],
    ["秋本 壮", "経営・営業", "", True, "1111"],
    ["福島 優美", "IPMD・トレンドリサーチ", "", True, "1111"],
    ["引地 瑞生", "IPMD・トレンドリサーチ", "", True, "1111"],
    ["齊藤 マミ", "IPMD・トレンドリサーチ", "", True, "1111"],
    ["川田 嘉悠", "ブランドPM・企画・EC", "", True, "1111"],
    ["小林 準基", "ブランドPM・企画・EC", "", True, "1111"],
    ["菊田 舞", "ブランドPM・企画・EC", "", True, "1111"],
    ["北園 佳樹", "ブランドPM・企画・EC", "", True, "1111"],
    ["藤本 有璃子", "クリエイティブ・デザイン", "fujimoto@konnekt-i.com", True, "1111"],
    ["高垣 ひなた", "クリエイティブ・デザイン", "", True, "1111"],
    ["植田 尋", "クリエイティブ・デザイン", "", True, "1111"],
    ["青木 和輝", "生産・物流・バックオフィス", "", True, "1111"],
    ["井道 桃香", "生産・物流・バックオフィス", "", True, "1111"],
    ["佐生 竜輔", "生産・物流・バックオフィス", "", True, "1111"],
]
for m in members:
    ws1.append(m)
style_header(ws1, 5)
ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 24
ws1.column_dimensions["C"].width = 26
ws1.column_dimensions["D"].width = 12
ws1.column_dimensions["E"].width = 10

# ============ タブ2: 施策 ============
ws2 = wb.create_sheet("施策")
ws2.append(["施策ID", "施策名", "月", "詳細", "有効フラグ"])

initiatives = [
    ["I001", "Claude × Shopify連携設定", "2026-05", "Shopifyコネクタを社員アカウントで接続", True],
    ["I002", "Claude議事録自動化（CircleBack）", "2026-05", "Google Meet→自動議事録", True],
    ["I003", "プロンプト集50選 読了", "2026-05", "社内ナレッジページを通読", True],
]
for i in initiatives:
    ws2.append(i)
style_header(ws2, 5)
ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 32
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 40
ws2.column_dimensions["E"].width = 12

# ============ タブ3: 回答 ============
ws3 = wb.create_sheet("回答")
ws3.append([
    "タイムスタンプ", "名前", "部署",
    "Q1頻度", "Q2場面JSON", "Q3効率化%", "Q4成長",
    "Q5目標", "Q6共有", "Q7BP",
    "施策実施JSON", "できたこと", "コメント", "スコア"
])
style_header(ws3, 14)
for col, w in zip("ABCDEFGHIJKLMN", [20, 14, 22, 12, 30, 12, 10, 14, 10, 30, 30, 30, 30, 10]):
    ws3.column_dimensions[col].width = w

# ============ タブ4: 設定 ============
ws4 = wb.create_sheet("設定")
ws4.append(["キー", "値"])
settings = [
    ["人件費単価_時給", 5000],
    ["AI月額コスト", 50000],
    ["ADMIN_PASSWORD", "admin2026"],
    ["EXEC_PASSWORD", "exec2026"],
    ["SLACK_RECIPIENTS", ""],
]
for s in settings:
    ws4.append(s)
style_header(ws4, 2)
ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 40

# ============ タブ5: Q5選択肢 ============
ws5 = wb.create_sheet("Q5選択肢")
ws5.append(["値", "絵文字", "ラベル", "スコア", "有効フラグ"])
q5_options = [
    ["個人向上", "📚", "個人向上（自分のスキル磨く）", 10, True],
    ["チーム巻き込み", "🤝", "チーム巻き込み（広めていく）", 10, True],
    ["新ツール開拓", "🆕", "新ツール開拓（新しい挑戦）", 10, True],
    ["現状維持", "🌿", "現状維持でいい", 3, True],
]
for o in q5_options:
    ws5.append(o)
style_header(ws5, 5)
ws5.column_dimensions["A"].width = 18
ws5.column_dimensions["B"].width = 8
ws5.column_dimensions["C"].width = 36
ws5.column_dimensions["D"].width = 10
ws5.column_dimensions["E"].width = 12

# 保存
out = Path.home() / "Downloads" / "KONNEKT_AIPulse_データ.xlsx"
wb.save(out)
print(f"✅ 保存しました: {out}")
print(f"   タブ数: {len(wb.sheetnames)}")
print(f"   タブ名: {', '.join(wb.sheetnames)}")
